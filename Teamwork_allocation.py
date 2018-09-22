""" 
Program name      : Teamwork_allocation.py
Author            : Francesc Box
Date created      : September 11th, 2018

Purpose           : Generate an allocation report out of a Teamwork time
                    exported in excel for a certain period

Revision History  :
Date    | Author        | Ref   
======== =============== ===================
"""
################################## IMPORTS ###################################
import os
import openpyxl # To process excel files
from openpyxl.styles import Alignment, Font           # To style the cells
from decimal import Decimal, ROUND_DOWN               # To calculate percentages
import pyexcel as p                                   # To copy xls to xlsx
from pathlib import Path                              # To check if file exists
################################## CONSTANTS ###################################
FILE_OUT='Allocation Report.xlsX'
NEW_SHEET='Allocation %'
NEW_FTE='Project FTE'
CPROJECT=0                                 # Excel column containing the project
CWHO=0                                     # Excel column containing the consultant
CHOURS=0                                   # Excel column containing the hours logged
ROWS=0                                     # Number of rows in the excel
COLUMNS=0                                  # Number of columns in the excel
################################ DICTS & LISTS ################################# 
Total_hours_consultant={}                  # Total hours per consultant
Total_hours_cons_proj={}                   # Total hours per consultant and project
FTE_project={}                             # FTE per project
################################## FUNCTIONS ###################################
def open_excel(excel_name):
    # Opens the excel in xls format if exists, converts it to xlsx  
    # Check by column header which contain project, consultant and hours
    print('Opening and converting Excel file...')
    # If file exists, opens the excel and converts it to xlsx
    if Path(excel_name+".xls").is_file():   
        exists=True
        # Converts to xlsx 
        p.save_book_as(file_name=excel_name+".xls", dest_file_name=FILE_OUT)
        # Open the excel file and keep some values for later
        wb = openpyxl.load_workbook(FILE_OUT)
        sheet=wb['Overview']
        ROWS = sheet.max_row + 1        # Need to increase 1 as it starts in 0
        COLUMNS = sheet.max_column
        # Find out columns for Project, Consultant, and hours
        for i in range (1, COLUMNS):
            cell_value = sheet.cell(row=1, column=i).value
            if cell_value == "Project":
                CPROJECT = i
                continue
            elif cell_value == "Who":
                CWHO = i
                continue
            elif cell_value == "Decimal Hours":
                CHOURS = i
                continue
            elif cell_value == "Date":
                CDATE = i
                continue
        return(wb, sheet, exists)
    else:
        exists=False
        print('File {} does not exist !!!!'.format(excel_name+".xls"))
        return("", "", exists)


def accumulate_hours():
    # Process the list with all the time entries and accumulates hours for later
    #   calculation
    for timesheet in TimeSheets:
        proj=timesheet[0]
        cons=timesheet[1]
        hours=timesheet[2]
        # Hours per consultant
        try:   
            Total_hours_consultant[cons] += hours   
        except KeyError:
            Total_hours_consultant[cons] = hours
        # Hours per consultant/project
        if cons and proj and (cons,proj) in Total_hours_cons_proj.keys(): 
            Total_hours_cons_proj[(cons,proj)] += hours   
        else:
            Total_hours_cons_proj[(cons,proj)] = hours


def generate_allocation():
    # Generates a new tab in the excel with the allocation results
    cur_row=1
   
    if NEW_SHEET in wb.sheetnames: 
        wb.remove(wb[NEW_SHEET])
    wb.create_sheet(index=1, title=NEW_SHEET)
    sheet=wb[NEW_SHEET]
    
    # Header text
    sheet.cell(row=cur_row, column=1).value = "Consultant"
    sheet.cell(row=cur_row, column=2).value = "Total Hours"
    sheet.cell(row=cur_row, column=3).value = "Project"
    sheet.cell(row=cur_row, column=4).value = "Hours"
    sheet.cell(row=cur_row, column=5).value = "%"
    
    # Header Style
    for column in "ABCDE":
        cell=sheet[column+"1"]
        cell.font = Font(bold=True, italic=True)
        if column in "BDE":
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Heather Width
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 37
    sheet.column_dimensions['D'].width = 8
    sheet.column_dimensions['E'].width = 6

    cur_row += 1
    # Generate rows
    for item in sorted(Total_hours_cons_proj.items()):
        cons=item[0][0]
        proj=item[0][1]
        hours_proj= item[1]

        # Retrieve total hours done by consultant
        hours_consultant = Total_hours_consultant[cons]                   
        perc_hours = hours_proj * 100 / hours_consultant

        if perc_hours > 0:
            # Accumulate for later FTE
            if proj in FTE_project.keys():
                FTE_project[proj] += perc_hours
            else:
                FTE_project[proj] = perc_hours

            sheet.cell(row=cur_row, column=1).value = cons                          # Consultant
            sheet.cell(row=cur_row, column=2).value = hours_consultant              # Total hours consultant
            sheet.cell(row=cur_row, column=3).value = proj                          # Project
            sheet.cell(row=cur_row, column=4).value = hours_proj                    # Total hours consultant/project
            sheet.cell(row=cur_row, column=5).value = perc_hours                    # % Allocation

            # Row Style
            for column in "BDE":
                cell=sheet[column+str(i)]
                cell.alignment = Alignment(horizontal="center", vertical="center")

            cur_row += 1


def generate_FTE():
    # Generate a new sheet in the excel with the FTE per project
    if NEW_FTE in wb.sheetnames: 
        wb.remove(wb[NEW_FTE])
    wb.create_sheet(index=2, title=NEW_FTE)
    sheet=wb[NEW_FTE]

    # Header text
    sheet.cell(row=1, column=1).value = "Project"
    sheet.cell(row=1, column=2).value = "FTE"
    
    # Header Style
    cell=sheet["A1"]
    cell.font = Font(bold=True, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell=sheet["B1"]
    cell.font = Font(bold=True, italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
            
    # Heather Width
    sheet.column_dimensions['A'].width = 37
    sheet.column_dimensions['B'].width = 8

    # Generate rows
    i=2
    for item in sorted(FTE_project.items()):
        proj=item[0]
        FTE= item[1]
        
        sheet.cell(row=i, column=1).value = proj      # Project                    
        
        perc_hours=Decimal(FTE/100)
        perc_hours = round(perc_hours, 2)
        if perc_hours == 0:  # If too short, we use full decimals.
            perc_hours=float(FTE/100)

        sheet.cell(row=i, column=2).value = perc_hours       
        
        cell=sheet['B'+str(i)]
        cell.alignment = Alignment(horizontal="center", vertical="center")

        i += 1

##################################### MAIN #####################################
if __name__ == "__main__":
    os.system("cls" if os.name == "nt" else "clear")  # Clears the screen                     

    # Open Teamwork excel and convert to xlsx if needed
    wb, sheet, exists = open_excel('All Time Report')   

    if exists:
        print('Storing excel in memory')
        # Process the excel and store values in a list of lists
        TimeSheets=[]
        for i in range (2, ROWS):
            project = sheet.cell(row=i, column=CPROJECT).value
            consultant = sheet.cell(row=i, column=CWHO).value
            hours = float(sheet.cell(row=i, column=CHOURS).value)
            TimeSheets.append ([project, consultant, hours])

        print('Accumulating hours per consultant and project')
        accumulate_hours()

        print('Calculating allocation and generating new excel tab')
        generate_allocation()
        
        print('Calculating FTE and generating new excel tab')
        generate_FTE()

        print('Saving new file to disk')
        wb.save(FILE_OUT)

    print ("********** FINISHED")
